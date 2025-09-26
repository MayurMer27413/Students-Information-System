import openpyxl
from openpyxl import Workbook


# Sample data
headers = [
    'Enrollment Number', 'Name', 'Email ID', 'Contact Number', 'DOB', 
    'Program', 'Course', 'Specialisation', 'Father Name', 'Father Occupation', 'Profile Photo'
]

data = [
    ['2203051240127', 'John Doe', 'john@example.com', '9876543210', '01/01/2000', 'B.Tech', 'Computer Science', 'Software Engineering', 'Robert Doe', 'Businessman', 'https://randomuser.me/api/portraits/men/1.jpg'],
    ['2203051240088', 'Priya Shah', 'priya@example.com', '8765432109', '15/05/2001', 'B.Sc', 'Physics', 'Quantum Physics', 'Vikram Shah', 'Businessman', 'https://randomuser.me/api/portraits/women/2.jpg'],
    ['2203051240124', 'Arjun Patel', 'arjun@example.com', '7654321098', '22/07/1999', 'B.Tech', 'Artificial Intelligence', 'Machine Learning', 'Rajesh Patel', 'Businessman', 'https://randomuser.me/api/portraits/men/3.jpg'],
    ['2403051240022', 'Emma Wilson', 'emma@example.com', '6543210987', '30/11/2000', 'B.Com', 'Commerce', 'Finance', 'George Wilson', 'Businessman', 'https://randomuser.me/api/portraits/women/4.jpg'],
    ['2203051240125', 'Michael Chen', 'michael@example.com', '5432109876', '05/03/2001', 'B.Tech', 'Electronics', 'VLSI Design', 'Wei Chen', 'Engineer', 'https://randomuser.me/api/portraits/men/5.jpg'],
    ['2203051240126', 'Sophia Rodriguez', 'sophia@example.com', '4321098765', '18/09/2000', 'B.A.', 'Psychology', 'Clinical Psychology', 'Carlos Rodriguez', 'Doctor', 'https://randomuser.me/api/portraits/women/6.jpg'],
    ['2203051240128', 'Raj Kumar', 'raj@example.com', '3210987654', '27/02/1999', 'M.Tech', 'Computer Science', 'Data Science', 'Suresh Kumar', 'Professor', 'https://randomuser.me/api/portraits/men/7.jpg'],
    ['2203051240129', 'Aisha Khan', 'aisha@example.com', '2109876543', '12/12/2001', 'B.Sc', 'Chemistry', 'Organic Chemistry', 'Farhan Khan', 'Lawyer', 'https://randomuser.me/api/portraits/women/8.jpg'],
    ['2203051240130', 'David Smith', 'david@example.com', '1098765432', '08/06/2000', 'BBA', 'Business Administration', 'Marketing', 'James Smith', 'Accountant', 'https://randomuser.me/api/portraits/men/9.jpg'],
    ['2203051240131', 'Mei Lin', 'mei@example.com', '0987654321', '19/04/2001', 'B.Arch', 'Architecture', 'Urban Design', 'Zhang Lin', 'Architect', 'https://randomuser.me/api/portraits/women/10.jpg']
]

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write headers
for col_idx, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_idx).value = header

# Write data
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx).value = cell_value

# Save the workbook
workbook.save('students.xlsx')
print('Created students.xlsx with sample records.')