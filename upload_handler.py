import os
import openpyxl
from flask import Flask, request, redirect, url_for, flash, render_template
from werkzeug.utils import secure_filename

# Configuration
UPLOAD_FOLDER = 'static/images/profile_photos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
EXCEL_FILE = 'students.xlsx'
ID_COLUMN = 'Enrollment Number'  # column name in the excel which holds unique IDs
PHOTO_COLUMN = 'Profile Photo'  # column name for student profile photos

# Helper functions
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def handle_photo_upload(app):
    # Ensure the upload folder exists
    os.makedirs(os.path.join(app.root_path, UPLOAD_FOLDER), exist_ok=True)
    
    # Add the route to the Flask app
    @app.route('/upload_photo', methods=['GET', 'POST'])
    def upload_photo():
        if request.method == 'POST':
            # Check if the student ID is provided
            student_id = request.form.get('student_id', '').strip()
            if not student_id:
                flash('Please enter a student ID.')
                return redirect(url_for('upload_photo'))
            
            # Check if the post request has the file part
            if 'profile_photo' not in request.files:
                flash('No file part')
                return redirect(request.url)
            
            file = request.files['profile_photo']
            
            # If user does not select file, browser also
            # submit an empty part without filename
            if file.filename == '':
                flash('No selected file')
                return redirect(request.url)
            
            if file and allowed_file(file.filename):
                # Secure the filename and create a unique filename with student ID
                filename = secure_filename(file.filename)
                # Use student ID in the filename to make it unique
                filename = f"{student_id}_{filename}"
                file_path = os.path.join(app.root_path, UPLOAD_FOLDER, filename)
                
                # Save the file
                file.save(file_path)
                
                # Update the Excel file with the photo path
                try:
                    # Load the Excel file
                    workbook = openpyxl.load_workbook(EXCEL_FILE)
                    sheet = workbook.active
                    
                    # Get headers from the first row
                    headers = [cell.value for cell in sheet[1]]
                    
                    # Find the column indices for ID and photo
                    id_col_idx = headers.index(ID_COLUMN) + 1  # +1 because openpyxl is 1-indexed
                    photo_col_idx = headers.index(PHOTO_COLUMN) + 1
                    
                    # Find the student by ID
                    student_found = False
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                        if str(row[id_col_idx - 1]).strip() == student_id:
                            student_found = True
                            # Use relative path for web access
                            relative_path = os.path.join(UPLOAD_FOLDER, filename).replace('\\', '/')
                            # Update the photo path in the sheet
                            sheet.cell(row=row_idx, column=photo_col_idx).value = relative_path
                            break
                    
                    if not student_found:
                        flash('Student not found.')
                        return redirect(url_for('upload_photo'))
                    
                    # Save the updated workbook back to Excel
                    workbook.save(EXCEL_FILE)
                    
                    flash('Profile photo uploaded successfully!')
                    return redirect(url_for('index'))
                    
                except Exception as e:
                    flash(f'Error updating Excel file: {str(e)}')
                    return redirect(url_for('upload_photo'))
            else:
                flash('File type not allowed. Please upload an image file (png, jpg, jpeg, gif).')
                return redirect(request.url)
        
        # If GET request, show the upload form
        return render_template('upload_photo.html')

# This function should be imported and called in app.py
# Example: from upload_handler import handle_photo_upload
#          handle_photo_upload(app)